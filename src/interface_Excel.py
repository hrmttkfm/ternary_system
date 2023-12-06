import openpyxl

class InterfaceExcel:
    def __init__(self, filepath: str):
        self.filepath = filepath # エクセルファイルのパス
        self.__vars = {
        # インターフェース側から設定可能なパラメータ群
        "Air Gap Length"               : {"val": None, "pos": "D3" , "init_val": None}, # m
        "Air Gap Diameter"             : {"val": None, "pos": "D4" , "init_val": None}, # m 
        "Air Gap Temperature"          : {"val": None, "pos": "I10", "init_val": None}, # ℃
        "Dope Polymer Fraction"        : {"val": None, "pos": "N5" , "init_val": None}, # wt-ratio ← 添加剤を含むこともある
        "Dope Solvent Fraction"        : {"val": None, "pos": "N4" , "init_val": None}, # wt-ratio 
        "Dope Non-Solvent Fraction"    : {"val": None, "pos": "N3" , "init_val": None}, # wt-ratio
        "Coagulation Bath Temperature" : {"val": None, "pos": "I11", "init_val": None}, # ℃
        "Spinning Speed"               : {"val": None, "pos": "D6" , "init_val": None}, # m/s
        "Fiber Inner Diameter"         : {"val": None, "pos": "D9" , "init_val": None}, # m
        "Fiber Outer Diameter"         : {"val": None, "pos": "D8" , "init_val": None}, # m
        # エクセル計算で得られる値の群                 
        "Time To Pass"                 : {"val": None, "pos": "B22", "init_val": None}, # s
        "Phi"                          : {"val": None, "pos": "G22", "init_val": None}, # - ← 無次元水分重量分率
        "Non-Solvent Fraction"         : {"val": None, "pos": "K22", "init_val": None}
        }                                

    def openFile(self, filepath: str = None):
        if not isinstance(filepath, type(None)):
            self.filepath = filepath
        self.wb = openpyxl.load_workbook(self.filepath, data_only = True, keep_vba = True)
        self.ws = self.wb["計算用"]
        self.readInitValsFromSheet()

    def readInitValsFromSheet(self):
        for key in self.__vars.keys():
            val = self.ws[self.__vars[key]["pos"]].value
            # print(val)
            self.setInitVal(key, val)

    def closeFile(self):
        # self.writeInitValsToSheet()
        self.wb.close()

    def writeInitValsToSheet(self):
        for key in self.__vars.keys():
            if not key in ["Time To Pass", "Phi", "Non-Solvent Fraction"]:
                self.ws[self.__vars[key]["pos"]].value = self.__vars[key]["init_val"]
                self.wb.save(self.filepath)

    def setVal(self, var_name: str, val: float):
        """変数に値を設定する

        Args:
            var_name (str): 変数の名前(self.vars辞書におけるキーの名前)
            val (float): 変数の値
        """
        # assert(type(val) == float or type(val) == int)
        try:
            self.__vars[var_name]["val"] = val
        except:
            print(f"変数{var_name}には対応していません.")

    def setInitVal(self, var_name: str, val: float):
        """変数に値を設定する

        Args:
            var_name (str): 変数の名前(self.vars辞書におけるキーの名前)
            val (float): 変数の値
        """
        # assert(type(val) == float or type(val) == int)
        try:
            self.__vars[var_name]["init_val"] = val
        except:
            print(f"変数{var_name}には対応していません.")

    def writeValToSheet(self, var_name: str):
        """変数の値をExcelに書き込む

        Args:
            var_name (str): 変数の名前(self.vars辞書におけるキーの名前)
        """
        if not var_name in ["Time To Pass", "Phi", "Non-Solvent Fraction"]:
            self.ws[self.__vars[var_name]["pos"]].value = self.__vars[var_name]["val"]
            self.wb.save(self.filepath)
        else:
            print(f"変数{var_name}の書き込みは許可されていません.")

    def readValFromSheet(self, var_name: str):
        """Excelから変数の値を読み込む

        Args:
            var_name (str): 変数の名前(self.vars辞書におけるキーの名前)
        """
        self.setVal(var_name, self.ws[self.__vars[var_name]["pos"]].value)

    def getVal(self, var_name: str):
        """変数の値を取り出す

        Args:
            var_name (str): 変数の名前(self.vars辞書におけるキーの名前)
        """
        try:
            if not isinstance(self.__vars[var_name]["val"], type(None)):
                return self.__vars[var_name]["val"]
            else:
                print(f"変数{var_name}には値が設定されていません.")
        except:
            print(f"変数{var_name}には対応していません.")


if __name__ == "__main__":
    ei = InterfaceExcel(f"src\VIPS吸湿量計算.xlsm")
    ei.openFile()
    ei.setVal("Air Gap Length", 100)
    ei.writeValToSheet("Air Gap Length")
    ei.closeFile()